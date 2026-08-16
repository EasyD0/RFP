# 数据结构
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, overload

from MyPyLib.LogSet import logSetUp
from clang.cindex import Cursor
from openpyxl import load_workbook
from warnings import deprecated

logger = logSetUp(__name__)
code_text = {
    "file_name": "文件名",
    "func_name": "函数名",
    "code_line": "代码行",
    "rule_name": "规则名称",
    "rule_code": "违反码",
    "pro_des": "当前问题描述",
    "is_false_alarm": "真实/误报问题",
    "history_reason": "历史误报原因",
    "history_problem": "历史相同误报问题",
    "rule_ref": "规则引用",
    "level": "问题等级",
}


class CodePos:
    """
    一组代码行的代码位置, 包含的路径都是相对路径
    token 中不含空格
    """

    def __init__(
        self, path: Path, line: int, token: str = "", proj_dir: Path | None = None
    ):
        if proj_dir:
            self.path = (proj_dir / path).resolve()
        else:
            self.path = path

        self.line = line
        self.token = token.strip()  # 代码的文本

    def __contains__(self, item: str) -> bool:
        return item in self.token

    @classmethod
    def from_2rows(cls, rows: str, proj_dir: Path | None = None) -> "CodePos":
        """
        从一组代码行生成对象
        :param rows: 一组代码行的文本
        :param proj_dir:
        :return:
        """
        texts = rows.splitlines()
        if len(texts) == 2:
            p, token = texts
        elif len(texts) == 1:
            p: str = texts[0]
            token = ""
        else:
            raise ValueError(rows)

        path_str, line_str = p.rsplit(":", 1)
        return cls(
            path=Path(path_str.strip()),
            line=int(line_str.strip()),
            token=token.strip(),
            proj_dir=proj_dir,
        )

    def __lt__(self, other: "CodePos") -> bool:
        if not isinstance(other, CodePos):
            raise TypeError
        if self.path != other.path:
            return str(self.path) < str(other.path)
        if self.line != other.line:
            return self.line < other.line
        return self.token < other.token

    def __eq__(self, other: "CodePos") -> bool:
        if not isinstance(other, CodePos):
            return False
        return (
            self.path == other.path
            and self.line == other.line
            and self.token == other.token
        )

    @classmethod
    def from_cusor(cls, cursor: Cursor) -> "CodePos":
        line = cursor.extent.start.line
        file_name = str(cursor.extent.start.file)
        with open(file_name, "r", encoding="utf-8", errors="replace") as f:
            token = f.readlines()[line - 1]
        return cls(Path(file_name).resolve(), line, token)


class RuleName:
    def __init__(self, raw_text: str | None):
        self.raw_text: str = raw_text or ""
        self.token: str = ""
        self.__post_init__()

    def __post_init__(self):
        if not ":" in self.raw_text:
            return
        self.token = self.raw_text.split(":")[-1].strip()

    def __contains__(self, item) -> bool:
        if isinstance(item, str):
            return item in self.raw_text
        return False


class CodeLine:
    """
    内部含有的路径都是相对路径
    """

    def __init__(self, raw_text: str | None):
        self.raw_text: str = raw_text or ""
        self._lines: list[CodePos] = []
        self.__post_init__()

    def __eq__(self, other: "CodeLine"):
        if not isinstance(other, CodeLine):
            return False
        return self.raw_text == other.raw_text

    def __lt__(self, other: "CodeLine") -> bool:
        if not isinstance(other, CodeLine):
            raise TypeError
        if self == other:
            return False
        if not self.__len__() or not other.__len__():
            return len(self) < len(other)

        if self[0] != other[0]:
            return self[0] < other[0]

        if len(self) != len(other):
            return len(self) < len(other)
        else:
            for i in range(1, len(self)):
                if self[i] != other[i]:
                    return self[i] < other[i]
        raise AssertionError

    def __post_init__(self):
        if not self.raw_text:
            return
        # raw_texts = self.raw_text.splitlines()
        # for i in range(len(raw_texts)):
        #     if i % 2 == 0:
        #         path, line = raw_texts[i].split(":")
        #         self._lines.append(CodePos(Path(path), int(line), raw_texts[i + 1]))
        for two_row in self.raw_text.split("\n\n"):
            code_pos = CodePos.from_2rows(two_row)
            self._lines.append(code_pos)

    @classmethod
    def from_rawtext(cls, text) -> "CodeLine":
        res = cls(text)
        res.__post_init__()
        return res

    def __len__(self):
        return len(self._lines)

    def __iter__(self):
        return iter(self._lines)

    def __str__(self):
        return self.raw_text

    @overload
    def __getitem__(self, idx: int) -> CodePos: ...

    @overload
    def __getitem__(self, idx: slice) -> list[CodePos]: ...

    def __getitem__(self, idx) -> CodePos | list[CodePos]:
        return self._lines[idx]


@dataclass
class Problem:
    """
    问题数据
    """

    file_name: str = ""  # 文件名
    func_name: str = ""  # 函数名
    code_line: CodeLine = None  # 代码行
    rule_name: RuleName = None  # "规则名称"
    rule_code: str = ""  # "违反码"
    pro_des: str = ""  # "当前问题描述"
    is_false_alarm: bool = False  # "真实/误报问题"
    history_reason: str = ""  # "历史误报原因"
    history_problem: str = ""  # "历史相同误报问题"
    rule_ref: str = ""  # "规则引用"
    level: str = ""  # "问题等级"

    def __lt__(self, other: "Problem") -> bool:
        if not isinstance(other, Problem):
            raise TypeError

        if self.is_false_alarm != other.is_false_alarm:
            # 先按是否误报排序
            return other.is_false_alarm
        if self.level != other.level:
            return self.level < other.level
        if self.rule_code != other.rule_code:
            return self.rule_code < other.rule_code
        if self.code_line and other.code_line:
            if not self.code_line != other.code_line:
                return self.code_line < other.code_line
        if self.file_name != other.file_name:
            return str(self.file_name) < str(other.file_name)

        return str(self.func_name) < str(other.func_name)

    def __eq__(self, other: "Problem") -> bool:
        if not isinstance(other, Problem):
            raise TypeError

        return (
            self.is_false_alarm == other.is_false_alarm
            and self.level == other.level
            and self.rule_code == other.rule_code
            and self.file_name == other.file_name
            and self.func_name == other.func_name
            and self.code_line == other.code_line
        )

    def set_false(self):
        """
        设为误报
        :return:
        """
        self.is_false_alarm = True

    def clear_false(self):
        self.is_false_alarm = False

    def file_path1(self, proj_dir: Path | None = None) -> Path:
        """
        返回代码行里的第一个路径
        :param proj_dir:
        :return:
        """
        try:
            if proj_dir:
                return proj_dir / self.code_line[0].path
            else:
                return self.code_line[0].path
        except IndexError:
            logger.error("数组越界")
            return Path(self.file_name)

    def file_path2(self, proj_dir: Path | None = None) -> Path:
        """
        返回代码行里的第一个路径
        :param proj_dir:
        :return:

        """
        try:
            if proj_dir:
                return proj_dir / self.code_line[1].path
            else:
                return self.code_line[1].path
        except IndexError:
            logger.error("数组越界")
            return Path(self.file_name)

    def to_dict(self) -> dict:
        """
        转为字典
        """
        return {
            "文件名": self.file_name,
            "函数名": self.func_name,
            "代码行": self.code_line.raw_text,
            "规则名称": self.rule_name.raw_text,
            "违反码": self.rule_code,
            "当前问题描述": self.pro_des,
            "真实/误报问题": "误报" if self.is_false_alarm else "真实",
            "历史误报原因": self.history_reason,
            "历史相同误报问题": self.history_problem,
            "规则引用": self.rule_ref,
            "问题等级": self.level,
        }

    @classmethod
    def from_dict(cls, data: dict[str, str]) -> "Problem":
        res = cls()
        for k in res.__dict__.keys():
            if k in {"code_line", "rule_name"}:
                continue

            if k in data or code_text[k] in data:
                res.__dict__[k] = data.get(k) or data.get(code_text[k])

        res.code_line = CodeLine(
            data.get("code_line") or data.get(code_text["code_line"])
        )

        res.rule_name = RuleName(
            data.get("rule_name") or data.get(code_text["rule_name"])
        )
        return res

    @deprecated("已弃用")
    def apply_checker(self, checker: Callable[["Problem"], bool]) -> "Problem":
        """
        应用检查
        :param checker: 将返回对 is_false_alarm 的标记
        :return:
        """
        self.is_false_alarm = checker(self)
        return self


def load_from_table(file: Path) -> list[Problem]:
    """
    从表格文件中读取

    表头为
    序号	文件名	函数名	代码行	规则名称	违反码	当前问题描述	真实/误报问题	历史误报原因	历史相同误报问题	规则引用	问题等级	处理方式(修改/不修改)	处理情况说明	确认人

    :return:
    """
    # 列索引映射（从1开始）
    col_map = {
        "序号": 1,
        "文件名": 2,
        "函数名": 3,
        "代码行": 4,
        "规则名称": 5,
        "违反码": 6,
        "当前问题描述": 7,
        "真实/误报问题": 8,
        "历史误报原因": 9,
        "历史相同误报问题": 10,
        "规则引用": 11,
        "问题等级": 12,
        "处理方式(修改/不修改)": 13,
        "处理情况说明": 14,
        "确认人": 15,
    }

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    # 查找表头行
    header_row = None
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=10, max_col=15), start=1
    ):
        cell_values = [cell.value for cell in row]
        if "文件名" in cell_values and "函数名" in cell_values:
            header_row = row_idx
            break

    if header_row is None:
        logger.error(f"未找到表头行: {file}")
        return []

    problems: list[Problem] = []

    # 从表头下一行开始读取数据
    for row in ws.iter_rows(min_row=header_row + 1):
        # 检查第一列是否有数据（序号列）
        if not row[0].value:
            continue

        row_data = {
            col: row[col_idx - 1].value
            for col, col_idx in col_map.items()
            if col_idx <= len(row)
        }

        # 跳过无效行
        if not row_data.get("文件名"):
            continue

        # 转换 is_false_alarm
        is_false = row_data.get("真实/误报问题", "")
        if isinstance(is_false, str):
            is_false_alarm = "误报" in is_false
        else:
            is_false_alarm = bool(is_false)

        _get = lambda x: str(row_data.get(x) or "")
        # 构建 Problem 对象
        problem = Problem(
            file_name=_get("文件名"),
            func_name=_get("函数名"),
            code_line=CodeLine(_get("代码行")),
            rule_name=RuleName(_get("规则名称")),
            rule_code=_get("违反码"),
            pro_des=_get("当前问题描述"),
            is_false_alarm=is_false_alarm,
            history_reason=_get("历史误报原因"),
            history_problem=_get("历史相同误报问题"),
            rule_ref=_get("规则引用"),
            level=_get("问题等级"),
        )
        problems.append(problem)

    return problems


